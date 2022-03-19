import openpyxl

class AllTestsTestData:
    # Create a class/global variable to contain the test data for each one of the testCases.
    # We can define multiple testData variable for the same page.
    # Each defined testData variable can be used by different testCases.
    # in each testCase the class variable are calle with className.variableName convention.
    home_page_test_data_all = [
        {"environment": "TEST", "name": "Hector", "email": "TEST@gmail.com", "gender": "Male"},
        {"environment": "TEST", "name": "Ive", "email": "TEST@de.com", "gender": "Female"},
        {"environment": "TEST", "name": "Gisela", "email": "TEST@princess.com", "gender": "Female"},
        {"environment": "PRODUCTION", "name": "John", "email": "PRODUCTION@gmail.com", "gender": "Male"},
        {"environment": "PRODUCTION", "name": "Ive", "email": "PRODUCTION@de.com", "gender": "Female"},
        {"environment": "STAGING", "name": "Githa", "email": "STAGING@de.com", "gender": "Female"},
        {"environment": "STAGING", "name": "Akinsha", "email": "STAGING@de.com", "gender": "Female"}
         ]

    buy_product_test_data_1 = [{"product": "Nokia", "deliveryCountry": "France"},
                               {"product": "Blackberry", "deliveryCountry": "India"},
                               {"product": "apple", "deliveryCountry": "Russia"}]

    # To inherit and use any variable from a class we can use the notation/format: "className.variableName" in the
    # child class. But to use/inherit any method from a class we have to instantiate/create an object of the parent
    # class in the child class. But if we don't want to instantiate any object we can define the method as static and
    # then we can call the method as we do to inherit any variable, this is using the notation/
    # format: "className.methodName". To make the method static we use the decorator "@staticmethod" before the
    # method declaration, like below, also we WE HAVE TO REMOVE THE 'self' PARAMETER FROM THE METHOD DECLARATION. The
    # self parameter is only used in the declaration of 'non-static-method':
    @staticmethod
    def get_loging_test_data(environment="TEST"):

        dictionary_3 = {}

        # 2- load the Excel sheet, to a data-book object, using the load_workbook("excelsheetpath") method.
        hh_test_data_book = openpyxl.load_workbook(
            "C:\\Users\\ssshh\\PycharmProjects\\E2EFrameWork\\TestData\\test_data_Login_In_Excel.xlsx")

        # 3- create an 'active_sheet' object to point/link/access the active sheet in the work_book (hh_test_data_book)
        active_sheet = hh_test_data_book.active

        # another code for practicing:
        selected_users_list = []

        for i in range(1, active_sheet.max_row + 1):
            if active_sheet.cell(row=i, column=1).value == environment:
                for j in range(2, active_sheet.max_column + 1):
                    dictionary_3[active_sheet.cell(row=1, column=j).value] = \
                        str(active_sheet.cell(row=i, column=j).value)
                selected_users_list.append(dictionary_3)
                dictionary_3 = {}

        print("dictionary_3: ", dictionary_3)
        print("selected users list: ", selected_users_list)

        return selected_users_list  # to return the List containing the whole dictionary with all the ligin for
                                    # the environment requested.
        # return [dictionary_3]  # to return the complete dictionary with all the ligin for the environment requested.

