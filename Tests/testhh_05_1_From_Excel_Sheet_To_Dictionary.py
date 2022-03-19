# this program move data from an Excel spreadsheet to a python Dictionary data structure:
#

import openpyxl

# 1- define the dictionary variable
dictionary_1 = {}
dictionary_2 = {}
dictionary_3 = {}

# 2- load the Excel sheet, to a data-book object, using the load_workbook("excelsheetpath") method.
hh_test_data_book = openpyxl.load_workbook("C:\\Users\\ssshh\\PycharmProjects\\E2EFrameWork\\TestData\\test_data_excel.xlsx")

# 3- create an 'active_sheet' object to point/link/access the active sheet in the work_book (hh_test_data_book)
active_sheet = hh_test_data_book.active

# 4 To move the content of the whole spreadsheet to dictionary_1, we use two for....loop:
# assuming that row one has the headers of each column:
for i in range(2, active_sheet.max_row + 1):
    for j in range(1, active_sheet.max_column + 1):
       dictionary_1[active_sheet.cell(row=1, column=j).value + str(i-1)] = \
           str(active_sheet.cell(row=i, column=j).value)
    # Note:
    #     Added concatenation of str(i) to the 'key' name to avoid having the same key for each row and not
    #     over writing the previous content.
print("Dict_1: ", dictionary_1)

# to move to a dictionary only the cells for a row having an specific content in column # 1, but not including the column # 1.
for i in range(1, active_sheet.max_row + 1):
    if active_sheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, active_sheet.max_column + 1):  # start in 2 to avoid printing column # 1.
            dictionary_2[active_sheet.cell(row=1, column=j).value] = \
                str(active_sheet.cell(row=i, column=j).value)

print("Dict_2", dictionary_2)

# Below replace the content of the dictionary with the data for 'Testcase2' just need to change the if statement
# because the key has only 'active_sheet.cell(row=1, column=j).value':
for i in range(1, active_sheet.max_row + 1):
    if active_sheet.cell(row=i, column=1).value == "Testcase2":
        for j in range(2, active_sheet.max_column + 1):  # start in 2 to avoid printing column # 1.
            dictionary_2[active_sheet.cell(row=1, column=j).value] = \
                str(active_sheet.cell(row=i, column=j).value)
print("Dict_2", dictionary_2)

# another code for practicing:
test_case = "Testcase4"
for i in range(1, active_sheet.max_row + 1):
    print("value in cell = " + active_sheet.cell(row=i, column=1).value)
    if active_sheet.cell(row=i, column=1).value == test_case:
        print("value: ", active_sheet.cell(row=i, column=1).value)
        for j in range(2, active_sheet.max_column + 1):
            dictionary_3[active_sheet.cell(row=1, column=j).value] = \
                str(active_sheet.cell(row=i, column=j).value)
print("new dictionary practice: ", dictionary_3)

# Create a Data type List with elements containing Dictionary structure for each test case in the spreadsheet:
all_testcases_list = []
dict_4 = {}
for i in range(2, active_sheet.max_row + 1):
    for j in range(2, active_sheet.max_column + 1):
        dict_4[active_sheet.cell(row=1, column=j).value] = \
            str(active_sheet.cell(row=i, column=j).value)

    print("Dict_4:", dict_4)
    all_testcases_list.append(dict_4)
    dict_4 = {}

print("all testcases: ", all_testcases_list)