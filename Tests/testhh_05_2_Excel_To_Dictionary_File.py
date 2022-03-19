# this program move data from an Excel spreadsheet to a python Dictionary data structure:
#

import openpyxl

# 1- define the dictionary variable
dictionary_3 = {}

# 2- load the Excel sheet, to a data-book object, using the load_workbook("excelsheetpath") method.
hh_test_data_book = openpyxl.load_workbook("C:\\Users\\ssshh\\PycharmProjects\\E2EFrameWork\\TestData\\test_data_Login_In_Excel.xlsx")

# 3- create an 'active_sheet' object to point/link/access the active sheet in the work_book (hh_test_data_book)
active_sheet = hh_test_data_book.active

# another code for practicing:
selected_users_list = []

environment = "TEST"
for i in range(1, active_sheet.max_row + 1):
    if active_sheet.cell(row=i, column=1).value == environment:
        for j in range(2, active_sheet.max_column + 1):
            dictionary_3[active_sheet.cell(row=1, column=j).value] = \
                str(active_sheet.cell(row=i, column=j).value)
        selected_users_list.append(dictionary_3)
        dictionary_3 = {}
print("selected users list: ", selected_users_list)

with open('C:/Users/ssshh/PycharmProjects/E2EFrameWork/TestData/usersLogin.py', 'w') as output_file:
    output_file.write("class TestLoginData : ")
    output_file.write("selected_users = " + str(selected_users_list))

