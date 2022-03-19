# this program introduces the use of Excel spreadsheet how to Read and write to and from it.
# We need the openpyxl Python library to read and write to Excel 2010 xlsx/xlsm... files.
# to install the openpyxl package we have to issue the command: pip install openpyxl ,in our local system.
# make sure that the openpyxl is included in the interpreter where we have installed the Python, this is:
#       C:\Users\ssshh\AppData\Local\Programs\Python\Python39\python.exe
# 1- import the openpyxl library

import openpyxl

# 2- load the Excel sheet, to a data-book object, using the load_workbook("excelsheetpath") method.
hh_test_data_book = openpyxl.load_workbook("C:\\Users\\ssshh\\PycharmProjects\\E2EFrameWork\\TestData\\test_data_excel.xlsx")

# 3- create an 'active_sheet' object to point/link/access the active sheet in the work_book (hh_test_data_book)
active_sheet = hh_test_data_book.active

# 3- the READ/GET the data from a column in the active_sheet we have to use the 'value' method of the cell object:
#    sheet.cell(row=r#, column=c#).value ..
#    To access the value in cell(1, 2) row 1 column 2.
cell_r1_c2_value = active_sheet.cell(row=1, column=2).value
print("cell_r1_c2_value:", cell_r1_c2_value)

# using the object representing the cell(1,2):
cell_r1_c2_value_object = active_sheet.cell(row=1, column=2)
print("using the object to print the its value: cell_r1_c2_value_object.value:", cell_r1_c2_value_object.value)

# 4- To WRITE/UPDATE the value/content of ta sale we use the same 'value' method of the cell object, but in the
# left part of the equation: sheet.cell(row=r, column=c).value = "new_content"
active_sheet.cell(row=2, column=2).value = "Hector"
# now printing to double check:
print("new content in Cell2,2 is: ", active_sheet.cell(row=2, column=2).value)

# 5- to know the current number of rows and column in the active sheet we use the sheet's methods: max_row, max_column:
#    sheet.max_row returns the value 2 if the active sheet is empty.
#    sheet.max_column returns the value 2 if the active sheet is empty.
print("Number or rows in sheet: ", active_sheet.max_row)
print("number of columns in the sheet: ", active_sheet.max_column)

# the sheet.min_row give the value 1 if sheet is populated or not populated.
# the sheet.min_column give the value 1 if sheet is populated or 2 if not populated.
print("min_row: ", active_sheet.min_row)
print("min_column: ", active_sheet.min_column)

# we can access the value in the cell by the column letter and the row number: A2, B3 .., sheet['A2'].value, sheet['B3'].value
print("Cell A3 value: ", active_sheet['A3'].value)
print("Cell B1 value: ", active_sheet['b1'].value)
# also using a variable for the row number:
index = 4
print("using 'x' as a variable for row active_sheet['A'+x].value: ", active_sheet['A'+str(index)].value)
# or using a variable for the column letter:
col_letter = "A"
print("using a variable for the column letter:", active_sheet[col_letter+"6"].value)

# 6 to display the title of the active sheet use the '.title' method.
print("active sheet name: ", active_sheet.title)

# 7 to print all the cells in column 1 we use a for...loop with a range(from, to +1), remember the 'to' in not
# inclusive this is why we have to add 1.
for i in range(1, active_sheet.max_row + 1):
    print("column=1, row="+str(i)+" = " + active_sheet.cell(row=i, column=1).value)

# To print all the cells in row 1 use another for....loop with range..
for j in range(1, active_sheet.max_column + 1):
    print("row=1, column="+str(j)+" = "+active_sheet.cell(row=1, column=j).value)

# To print the content of the whole spreadsheet, we use two for....loop:
for i in range(1, active_sheet.max_row + 1):
    for j in range(1, active_sheet.max_column + 1):
        print("cell("+str(i)+","+str(j)+") = "+active_sheet.cell(row=i, column=j).value)

# to print only the cells for a row having an specific content in column # 1, i.e 'Testcase3'
for i in range(1, active_sheet.max_row + 1):
    if active_sheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(1, active_sheet.max_column + 1):
            print("cell("+str(i)+","+str(j)+") = "+active_sheet.cell(row=i, column=j).value)
# to print only the cells for a row having an specific content in column # 1, but not including the column # 1.
for i in range(1, active_sheet.max_row + 1):
    if active_sheet.cell(row=i, column=1).value == "Testcase3":
        for j in range(2, active_sheet.max_column + 1):  # start in 2 to avoid printing column # 1.
            print("cell("+str(i)+","+str(j)+") = "+active_sheet.cell(row=i, column=j).value)